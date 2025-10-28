from django.core.management.base import BaseCommand
from nadicomdigital.models import Kursus, Peserta
import openpyxl
from openpyxl.styles import Font, Alignment
from django.db.models import Count
import os  

class Command(BaseCommand):
    help = 'Export data kursus dan peserta ke Excel'
    
    def add_arguments(self, parser):
        parser.add_argument('filename', type=str, help='Nama file output')
    
    def handle(self, *args, **options):
        try:
            filename = options['filename']
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # ✅ DELETE FILE LAMA JIKA ADA
            if os.path.exists(filename):
                os.remove(filename)
                self.stdout.write(self.style.WARNING(f'🗑️ File lama {filename} dihapus'))
            
            # Dapatkan data kursus dengan jumlah peserta (DATA TERKINI)
            kursus_list = Kursus.objects.annotate(
                jumlah_peserta=Count('peserta')
            ).all()
            
            if not kursus_list.exists():
                self.stdout.write(self.style.WARNING('⚠️ Tidak ada data kursus untuk di-export'))
                return
            
            # Buat workbook Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Kursus"
            
            # Header row untuk data kursus
            headers = [
                'ID', 'Tajuk Kursus', 'Tarikh', 'Masa Mula', 'Masa Tamat',
                'Lokasi', 'Kapasiti', 'Jumlah Peserta', 'Slot Tersedia',
                'Status', 'Created At', 'Updated At'
            ]
            
            # Apply styling untuk header
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data rows untuk kursus (DATA TERKINI)
            for row_num, kursus in enumerate(kursus_list, start=2):
                ws.cell(row=row_num, column=1, value=kursus.id)
                ws.cell(row=row_num, column=2, value=kursus.tajuk)
                ws.cell(row=row_num, column=3, value=kursus.tarikh.strftime('%Y-%m-%d'))
                ws.cell(row=row_num, column=4, value=kursus.masa_mula.strftime('%H:%M'))
                ws.cell(row=row_num, column=5, value=kursus.masa_tamat.strftime('%H:%M'))
                ws.cell(row=row_num, column=6, value=kursus.lokasi)
                ws.cell(row=row_num, column=7, value=kursus.kapasiti)
                ws.cell(row=row_num, column=8, value=kursus.jumlah_peserta)
                ws.cell(row=row_num, column=9, value=kursus.slot_tersedia())
                ws.cell(row=row_num, column=10, value=kursus.get_status_display())
                ws.cell(row=row_num, column=11, value=kursus.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=12, value=kursus.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Auto adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Buat sheet kedua untuk data peserta (DATA TERKINI)
            ws_peserta = wb.create_sheet(title="Data Peserta")
            
            # Header untuk data peserta
            headers_peserta = [
                'ID', 'Nama Peserta', 'No IC', 'Syarikat', 'No Telefon', 'Email',
                'Kursus ID', 'Tajuk Kursus', 'Dihantar Pada', 'Disahkan', 
                'Emel Dihantar', 'Tarikh Emel Dihantar'
            ]
            
            # Apply styling untuk header peserta
            for col_num, header in enumerate(headers_peserta, 1):
                cell = ws_peserta.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = openpyxl.styles.PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            
            # Data untuk peserta (DATA TERKINI)
            peserta_list = Peserta.objects.select_related('kursus').all()
            for row_num, peserta in enumerate(peserta_list, start=2):
                ws_peserta.cell(row=row_num, column=1, value=peserta.id)
                ws_peserta.cell(row=row_num, column=2, value=peserta.nama)
                ws_peserta.cell(row=row_num, column=3, value=peserta.no_ic)
                ws_peserta.cell(row=row_num, column=4, value=peserta.syarikat)
                ws_peserta.cell(row=row_num, column=5, value=peserta.no_telefon)
                ws_peserta.cell(row=row_num, column=6, value=peserta.email)
                ws_peserta.cell(row=row_num, column=7, value=peserta.kursus.id)
                ws_peserta.cell(row=row_num, column=8, value=peserta.kursus.tajuk)
                ws_peserta.cell(row=row_num, column=9, value=peserta.dihantar_pada.strftime('%Y-%m-%d %H:%M:%S'))
                ws_peserta.cell(row=row_num, column=10, value='Ya' if peserta.disahkan else 'Tidak')
                ws_peserta.cell(row=row_num, column=11, value='Ya' if peserta.emel_dihantar else 'Tidak')
                if peserta.tarikh_emel_dihantar:
                    ws_peserta.cell(row=row_num, column=12, value=peserta.tarikh_emel_dihantar.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    ws_peserta.cell(row=row_num, column=12, value='-')
            
            # Auto adjust column width untuk sheet peserta
            for column in ws_peserta.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_peserta.column_dimensions[column_letter].width = adjusted_width
            
            # Simpan file (FILE BARU DENGAN DATA TERKINI)
            wb.save(filename)
            self.stdout.write(self.style.SUCCESS(
                f'✅ Berhasil export {kursus_list.count()} kursus dan {peserta_list.count()} peserta ke {filename}'
            ))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error: {str(e)}'))