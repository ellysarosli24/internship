from django.shortcuts import render, redirect, get_object_or_404
from .models import Video, Service,  TeamMember, GalleryImage, BlogPost, BlogCategory, Kursus, Peserta
from django import forms
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import re
from django.views.generic import ListView
from django.views.generic import DetailView
from django.db import models
from django.db.models import Q, Count
from django.contrib import messages
from .forms import PendaftaranForm
from django.http import HttpResponse
from openpyxl import Workbook
from django.http import FileResponse
import os 
from .auto_sync import auto_sync_new_videos
from django.core.paginator import Paginator
import sendgrid
from sendgrid.helpers.mail import Mail
from django.conf import settings

def homepage(request):
    blog_posts = BlogPost.objects.filter(is_published=True).order_by('-published_date')[:3]
    return render(request, 'homepage.html', {'blog_posts': blog_posts})

def service_list(request):
    services = Service.objects.all()
    categories = Service.CATEGORY_CHOICES  # Pastikan ini diambil dari model
    return render(request, 'service_list.html', {
        'services': services,
        'categories': categories
    })




class ProfileView(ListView):
    template_name = 'profile.html'
    context_object_name = 'team_members'
    model = TeamMember
    
    def get_queryset(self):
        # Remove filter is_active and order by name since those fields don't exist
        return TeamMember.objects.all()  # Just get all team members
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # ✅ AMBIL GALLERY IMAGES DARI DATABASE
        gallery_images = GalleryImage.objects.filter(is_active=True).order_by('order')[:8]
        
        # ✅ HANTAR gallery_images BUKAN gallery static
        context.update({
            'gallery_images': gallery_images,  # ✅ INI YANG BETUL!
            'company_stats': [
                {'icon': 'calendar-alt', 'value': '8+', 'label': 'Tahun Pengalaman'},
                {'icon': 'users', 'value': '500+', 'label': 'Pelanggan'},
                {'icon': 'map-marked-alt', 'value': '10+', 'label': 'Negeri Diliputi'},
                {'icon': 'certificate', 'value': '100%', 'label': 'Fokus SQL'}
            ],
            'mission': "Memberikan penyelesaian perakaunan digital yang mudah, berkesan dan memenuhi keperluan khusus perniagaan Malaysia dengan sokongan berterusan.",
        })
        
        return context
    
class BlogListView(ListView):
    model = BlogPost
    template_name = 'blog_list.html'
    context_object_name = 'posts'
    paginate_by = 6
    
    def get_queryset(self):
        # Pastikan hanya post yang published dan ada published_date
        return BlogPost.objects.filter(
            is_published=True, 
            published_date__isnull=False
        ).order_by('-published_date')
    

class BlogDetailView(DetailView):
    model = BlogPost
    template_name = 'blog_detail.html'
    context_object_name = 'blog_post'
    
    def get_queryset(self):
        return BlogPost.objects.filter(is_published=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        blog_post = self.object
        
        # Tambah view count
        BlogPost.objects.filter(id=blog_post.id).update(views=models.F('views') + 1)
        blog_post.refresh_from_db()
        
        # Dapatkan post terkini
        context['recent_posts'] = BlogPost.objects.filter(
            is_published=True
        ).exclude(id=blog_post.id).order_by('-published_date')[:5]
        
        # PERBAIKI INI: Dapatkan SEMUA kategori dengan post count
        from django.db.models import Count
        categories = BlogCategory.objects.annotate(
            post_count=Count('blogpost', filter=models.Q(blogpost__is_published=True))
        ).order_by('name')  # HAPUS filter post_count__gt=0
        
        # DEBUG: Print semua kategori ke console
        print("DEBUG: All categories with post count:")
        for cat in categories:
            print(f"- {cat.name}: {cat.post_count} posts")
        
        context['categories'] = categories
        
        # Dapatkan post berkaitan
        context['related_posts'] = BlogPost.objects.filter(
            is_published=True,
            categories__in=blog_post.categories.all()
        ).exclude(id=blog_post.id).distinct()[:3]
        
        return context
    

class BlogCategoryView(ListView):
    model = BlogPost
    template_name = 'blog_list.html'  # Gunakan template yang sama dengan senarai blog
    context_object_name = 'posts'
    paginate_by = 6
    
    def get_queryset(self):
        # Dapatkan post berdasarkan kategori
        category_slug = self.kwargs['slug']
        return BlogPost.objects.filter(
            is_published=True,
            published_date__isnull=False,
            categories__slug=category_slug  # Filter mengikut kategori
        ).order_by('-published_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Tambah info kategori ke context
        category_slug = self.kwargs['slug']
        context['category'] = get_object_or_404(BlogCategory, slug=category_slug)
        context['current_category'] = category_slug  # Untuk highlight kategori aktif
        return context
    
    
def blog_search(request):
    query = request.GET.get('q', '').strip()
  
    print(f"DEBUG: Search query received: '{query}'")
    
    if query:
        posts = BlogPost.objects.filter(
            Q(title__icontains=query) | 
            Q(content__icontains=query) |
            Q(excerpt__icontains=query),
            is_published=True,
            published_date__isnull=False
        ).order_by('-published_date')
        
        print(f"DEBUG: Found {posts.count()} results for '{query}'")
    else:
        posts = BlogPost.objects.none()
        print("DEBUG: No query provided")
    
    # Dapatkan categories untuk sidebar
    categories = BlogCategory.objects.annotate(
        post_count=Count('blogpost', filter=Q(blogpost__is_published=True))
    ).filter(post_count__gt=0)
    
    context = {
        'posts': posts,
        'query': query,
        'title': f'Hasil Carian: {query}' if query else 'Carian Blog',
        'categories': categories
    }
    
    return render(request, 'blog_search.html', context)



def daftar_kursus_view(request):
    """
    View untuk halaman single page yang mengandungi kedua-dua 
    senarai kursus dan form pendaftaran
    """
    from datetime import date
    today = date.today()
    
    # TAMBAH FILTER INI SAHAJA:
    kursus_aktif = Kursus.objects.filter(
        status='AKTIF', 
        tarikh__gte=today  # Hanya kursus yang tarikhnya >= hari ini
    ).order_by('tarikh')
    
    # TAMBAH INI UNTUK TEMPLATE:
    for kursus in kursus_aktif:
        kursus.is_expired = False  # Semua kursus dalam query ini belum expired
    
    return render(request, 'daftar_kursus.html', {
        'kursus_list': kursus_aktif,
        'form': PendaftaranForm()
    })



def proses_daftar(request):
    """
    View untuk memproses form pendaftaran (POST request)
    """
    if request.method == 'POST':
        kursus_id = request.POST.get('kursus_id')
        
        if not kursus_id:
            messages.error(request, 'Sila pilih kursus terlebih dahulu.')
            return redirect('daftar_kursus')
            
        kursus = get_object_or_404(Kursus, id=kursus_id)
        
        # Semak jika masih ada slot
        if kursus.slot_tersedia() <= 0:
            messages.error(request, 'Maaf, kursus ini sudah penuh.')
            return redirect('daftar_kursus')
        
        form = PendaftaranForm(request.POST)
        if form.is_valid():
            # Semak jika email sudah wujud untuk kursus ini
            if Peserta.objects.filter(email=form.cleaned_data['email'], kursus=kursus).exists():
                messages.error(request, 'Email ini sudah didaftarkan untuk kursus ini.')
                return redirect('daftar_kursus')
            
            # Simpan data dengan commit=False untuk tambah kursus dulu
            peserta = form.save(commit=False)
            peserta.kursus = kursus
            peserta.save()
            
            # Hantar emel pengesahan
            try:
                subject = f"Pengesahan Pendaftaran Kursus: {kursus.tajuk}"
                message = f"""
                Terima kasih {peserta.nama} kerana mendaftar untuk kursus kami.

                Butiran Pendaftaran:
                Nama: {peserta.nama}
                No IC: {peserta.no_ic}
                Syarikat: {peserta.syarikat}
                No Telefon: {peserta.no_telefon}
                Email: {peserta.email}
                
                Kursus: {kursus.tajuk}
                Tarikh: {kursus.tarikh}
                Masa: {kursus.masa_mula} - {kursus.masa_tamat}
                Lokasi: {kursus.lokasi}

                Sila hadir 15 minit lebih awal untuk pendaftaran.
                Untuk sebarang pertanyaan, sila hubungi 09-7191081
.

                Harap maklum.
                Nadicom Digital Sdn Bhd
                """
                
                send_mail(
                    subject,
                    message.strip(),  # strip() untuk buang whitespace
                    settings.DEFAULT_FROM_EMAIL,
                    [peserta.email],
                    fail_silently=False,
                )
                
                # Update status email
                peserta.emel_dihantar = True
                peserta.tarikh_emel_dihantar = timezone.now()
                peserta.save()
                
                messages.success(request, 'Pendaftaran anda berjaya! Emel pengesahan telah dihantar.')
                return redirect('pendaftaran_berjaya')
                
            except Exception as e:
                # Log error untuk debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Gagal hantar email: {str(e)}")
                
                # Jika gagal hantar email, masih simpan rekod tetapi tunjuk warning
                messages.warning(request, 'Pendaftaran berjaya tetapi emel pengesahan gagal dihantar. Sila hubungi kami.')
                return redirect('pendaftaran_berjaya')
        else:
            # Form tidak valid, paparkan error
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return redirect('daftar_kursus')
    
    # Jika bukan POST request, redirect ke halaman daftar
    return redirect('daftar_kursus')

def pendaftaran_berjaya(request):
    """Halaman kejayaan selepas pendaftaran"""
    return render(request, 'pendaftaran_berjaya.html')


def contactus(request):
    return render(request, 'contactus.html')

@csrf_exempt
def validate_whatsapp_number(request):
    if request.method == 'POST':
        phone_number = request.POST.get('phone', '')
        
        # Validate Malaysian phone number format
        if re.match(r'^(\+?6?01)[0-46-9]-*[0-9]{7,8}$', phone_number):
            # Clean number: remove +, -, and spaces
            clean_number = re.sub(r'[\+\-\s]', '', phone_number)
            
            # Ensure it starts with 60 (Malaysia country code)
            if not clean_number.startswith('60'):
                if clean_number.startswith('0'):
                    clean_number = '60' + clean_number[1:]
                else:
                    clean_number = '60' + clean_number
            
            return JsonResponse({
                'valid': True,
                'clean_number': clean_number
            })
        else:
            return JsonResponse({
                'valid': False,
                'error': 'Format nombor tidak sah. Contoh: 0123456789 atau 60123456789'
            })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def export_kursus_excel(request):
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Kursus"
    
    # Get current data
    kursus_list = Kursus.objects.annotate(jumlah_peserta=Count('peserta')).all()
    
    # Header
    headers = ['ID', 'Tajuk Kursus', 'Tarikh', 'Masa Mula', 'Masa Tamat', 'Lokasi', 'Kapasiti', 'Jumlah Peserta', 'Slot Tersedia', 'Status']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Data
    for row_num, kursus in enumerate(kursus_list, start=2):
        ws.cell(row=row_num, column=1, value=kursus.id)
        ws.cell(row=row_num, column=2, value=kursus.tajuk)
        ws.cell(row=row_num, column=3, value=kursus.tarikh.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=4, value=kursus.masa_mula.strftime('%H:%M'))
        ws.cell(row=row_num, column=5, value=kursus.masa_tamat.strftime('%H:%M'))
        ws.cell(row=row_num, column=6, value=kursus.lokasi)
        ws.cell(row=row_num, column=7, value=kursus.kapasiti)
        ws.cell(row=row_num, column=8, value=kursus.jumlah_peserta)
        ws.cell(row=row_num, column=9, value=kursus.slot_tersedia())
        ws.cell(row=row_num, column=10, value=kursus.get_status_display())
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="kursus_terkini.xlsx"'
    wb.save(response)
    
    return response

def download_export(request, filename):
    file_path = os.path.join(settings.MEDIA_ROOT, filename)
    return FileResponse(open(file_path, 'rb'), as_attachment=True)



def video_list(request):
    """View untuk video list page dengan search, filter dan AUTO-SYNC"""
    
    # ✅ AUTO-SYNC: Check for new videos when users access the page
    sync_result = auto_sync_new_videos()
    
    category_filter = request.GET.get('category', 'all')
    search_query = request.GET.get('q', '')
    
    # ✅ UPDATE: Order by published_at (YouTube date) instead of created_at
    videos = Video.objects.all().order_by('-published_at')
    
    # Filter oleh category
    if category_filter != 'all':
        videos = videos.filter(category=category_filter)
    
    # Search oleh query
    if search_query:
        videos = videos.filter(
            models.Q(title__icontains=search_query) |
            models.Q(description__icontains=search_query)
        )
    
    # ✅ UPDATE: Add pagination untuk better performance
    paginator = Paginator(videos, 12)  # 12 videos per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = Video.VIDEO_CATEGORIES
    
    return render(request, 'video.html', {
        'videos': page_obj,  # ✅ UPDATE: guna page_obj instead of videos
        'categories': categories,
        'selected_category': category_filter,
        'search_query': search_query,
        'sync_result': sync_result,  # ✅ NEW: untuk debug/notification
    })
def video_detail(request, video_id):
    try:
        # ✅ AUTO-SYNC: Check for new videos when users access video detail
        auto_sync_new_videos()
        
        video = get_object_or_404(Video, id=video_id)
        
        # ✅ FIXED: GUNA REAL YOUTUBE VIEW COUNT & LOCAL TRACKING
        if hasattr(video, 'view_count'):
            # Method 1: Fetch real YouTube view count
            try:
                from .youtube_service import fetch_video_details
                youtube_data = fetch_video_details(video.youtube_id)
                
                if youtube_data and 'view_count' in youtube_data:
                    # Update dengan real YouTube view count
                    video.view_count = youtube_data['view_count']
                    print(f"✅ Updated with YouTube view count: {video.view_count}")
                else:
                    # Fallback: Increment local view count
                    video.view_count += 1
                    print(f"✅ Incremented local view count: {video.view_count}")
                    
            except Exception as e:
                # Fallback jika API fail
                video.view_count += 1
                print(f"⚠️ API failed, using local count: {video.view_count}")
            
            video.save(update_fields=['view_count'])
        else:
            # Fallback jika field belum ada
            print("❌ view_count field not available")
        
        # Get related videos
        related_videos = Video.objects.filter(
            category=video.category
        ).exclude(id=video.id)[:4]
        
        # ✅ UPDATE: Get recent videos for sidebar (5 videos terkini)
        recent_videos = Video.objects.all().order_by('-published_at')[:5]
        
        # ✅ UPDATE: Use published_at instead of created_at untuk consistency
        same_category_videos = Video.objects.filter(
            category=video.category
        ).exclude(id=video.id).order_by('published_at')
        
        # Get next video (newer video)
        next_video = same_category_videos.filter(
            published_at__gt=video.published_at  # ✅ UPDATE: guna published_at
        ).first()
        
        # If no newer video, get the oldest video
        if not next_video:
            next_video = same_category_videos.order_by('published_at').first()  # ✅ UPDATE
        
        # Get previous video (older video)
        previous_video = same_category_videos.filter(
            published_at__lt=video.published_at  # ✅ UPDATE: guna published_at
        ).first()
        
        # If no older video, get the newest video
        if not previous_video:
            previous_video = same_category_videos.order_by('-published_at').first()  # ✅ UPDATE
        
        return render(request, 'video_detail.html', {
            'video': video,
            'related_videos': related_videos,
            'recent_videos': recent_videos,  # ✅ NEW: untuk sidebar
            'next_video': next_video,
            'previous_video': previous_video,
            'categories': Video.VIDEO_CATEGORIES,  # ✅ NEW: untuk sidebar categories
        })
        
    except Exception as e:
        print(f"Error in video_detail: {str(e)}")
        return render(request, 'error.html', {'error': str(e)})